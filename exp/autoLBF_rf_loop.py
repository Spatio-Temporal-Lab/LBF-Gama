import copy
import time

import numpy as np
import pandas as pd
from openpyxl import Workbook
from sklearn.ensemble import RandomForestClassifier

import lib.bf_util
import lib.lgb_url

df_train = pd.read_csv('../dataset/url_train.csv')
df_test = pd.read_csv('../dataset/url_test.csv')
df_query = pd.read_csv('../dataset/url_query.csv')

train_urls = df_train['url']
test_urls = df_test['url']
query_urls = df_query['url']

X_train = df_train.drop(columns=['url', 'url_type']).values.astype(np.float32)
y_train = df_train['url_type'].values.astype(np.float32)
X_test = df_test.drop(columns=['url', 'url_type']).values.astype(np.float32)
y_test = df_test['url_type'].values.astype(np.float32)
X_query = df_query.drop(columns=['url', 'url_type']).values.astype(np.float32)
y_query = df_query['url_type'].values.astype(np.float32)

# 设置参数
params = {
    'n_estimators': 1,  # 每一轮训练的树的数量
    'random_state': 42,
    'max_leaf_nodes': 20,
    'warm_start': True,  # 启用 warm_start 以便在下一轮中继续训练
}

n_true = df_train[df_train['url_type'] == 1].shape[0] + df_test[df_test['url_type'] == 1].shape[0]
n_false = df_train[df_train['url_type'] == 0].shape[0] + df_test[df_test['url_type'] == 0].shape[0]
n_test = len(df_test)


def evaluate_thresholds(prediction_results, y_true, bf_bytes):
    sorted_indices = np.argsort(prediction_results)
    sorted_predictions = prediction_results[sorted_indices]
    sorted_true = y_true[sorted_indices]

    fp = n_false
    tp = 0
    best_thresh = 0
    best_fpr_lbf = 1.0

    unique_sorted_predictions, idx = np.unique(sorted_predictions, return_index=True)

    n = len(unique_sorted_predictions)
    for i in range(n):
        thresh = unique_sorted_predictions[i]

        if i < n - 1:
            count_1 = np.sum(sorted_true[idx[i]:idx[i + 1]])
            tp += count_1
            fp -= idx[i + 1] - idx[i] - count_1
        else:
            count_1 = np.sum(sorted_true[idx[i]:n])
            tp += count_1
            fp -= n - idx[i] - count_1

        bf_count = tp
        fpr_bf = lib.bf_util.get_fpr(bf_count, bf_bytes)
        fpr_lgb = fp / n_false
        fpr_lbf = fpr_lgb + (1 - fpr_lgb) * fpr_bf

        if fpr_lbf < best_fpr_lbf:
            best_thresh = thresh
            best_fpr_lbf = fpr_lbf

    print(f'best thresh = {best_thresh} and best fpr = {best_fpr_lbf}')
    return best_thresh, best_fpr_lbf


def rf_validate_url(model, X_train, y_train, train_urls, X_test, y_test, test_urls, threshold):
    prediction_results = model.predict_proba(X_train)[:, 1]  # 获取正类的预测概率
    binary_predictions = (prediction_results > threshold).astype(int)
    indices_train = [i for i in range(len(X_train)) if binary_predictions[i] == 0 and y_train[i] == 1]

    # 对测试集进行预测
    prediction_results = model.predict_proba(X_test)[:, 1]
    binary_predictions = (prediction_results > threshold).astype(int)
    indices_test = [i for i in range(len(X_test)) if binary_predictions[i] == 0 and y_test[i] == 1]

    # 提取对应的url
    urls_train = train_urls.iloc[indices_train].values
    urls_test = test_urls.iloc[indices_test].values

    # 返回合并后的url数组
    return np.concatenate((urls_train, urls_test), axis=0)


def rf_query_url(model, bloom_filter, X_query, y_query, query_urls, threshold, draw):
    fn = 0
    fp = 0
    cnt_ml = 0
    cnt_bf = 0
    total = len(X_query)
    print(f"query count = {total}")

    prediction_results = model.predict_proba(X_query)[:, 1]

    if draw:
        import matplotlib.pyplot as plt
        bins = np.arange(0, 1.1, 0.1)  # 生成 [0, 0.1, 0.2, ..., 1.0] 的区间
        plt.hist(prediction_results, bins=bins, edgecolor='black', alpha=0.7)
        plt.xlabel('Prediction value')
        plt.ylabel('Frequency')
        plt.title('Histogram of All Predictions')
        plt.xticks(bins)
        plt.grid(axis='y', alpha=0.75)
        plt.show()

    binary_predictions = (prediction_results > threshold).astype(int)

    print(prediction_results)
    print(binary_predictions)

    for i in range(total):
        true_label = y_query[i]
        url = query_urls[i]
        prediction = binary_predictions[i]
        if prediction == 1:
            if true_label == 0:
                fp = fp + 1
                cnt_ml = cnt_ml + 1
        else:
            if url in bloom_filter:
                if true_label == 0:
                    fp = fp + 1
                    cnt_bf = cnt_bf + 1
            else:
                if true_label == 1:
                    fn = fn + 1

    print(f"fp: {fp}")
    print(f"total: {total}")
    print(f"fpr: {float(fp) / total}")
    print(f"fnr: {float(fn) / total}")
    print(f"cnt_ml: {cnt_ml}")
    print(f"cnt_bf: {cnt_bf}")
    return float(fp) / total


start_time = time.perf_counter_ns()

best_rf = None
size = 200 * 1024
best_fpr = 1.0
best_threshold = 0.5
epoch_max = 20
wb = Workbook()
ws = wb.active
rf = RandomForestClassifier(**params)

for epoch_now in range(epoch_max):
    rf.fit(X_train, y_train)

    train_pred = rf.predict_proba(X_train)[:, 1]
    test_pred = rf.predict_proba(X_test)[:, 1]

    # 拼接预测结果
    all_predictions = np.concatenate([train_pred, test_pred])
    all_true_labels = np.concatenate([y_train, y_test])

    model_size = lib.lgb_url.lgb_get_model_size(rf)
    print(f'epoch now = {epoch_now}')
    print(f'model size = {model_size}')
    bf_bytes = size - model_size
    if bf_bytes <= 0:
        break

    best_thresh, best_fpr_lbf = evaluate_thresholds(all_predictions, all_true_labels, bf_bytes)

    print("模型在内存中所占用的大小（字节）:", model_size)

    data_negative = rf_validate_url(rf, X_train, y_train, train_urls, X_test, y_test, test_urls, best_thresh)
    print(f"{len(data_negative)} insert into bloom filter")
    bloom_filter = lib.lgb_url.create_bloom_filter(dataset=data_negative, bf_size=bf_bytes)
    # 访问布隆过滤器的 num_bits 属性
    num_bits = bloom_filter.num_bits
    # 将比特位转换为字节（8 bits = 1 byte）
    memory_in_bytes = num_bits / 8
    print("memory of bloom filter: ", memory_in_bytes)
    print("memory of learned model: ", model_size)
    fpr = rf_query_url(rf, bloom_filter, X_query, y_query, query_urls, best_thresh, False)
    ws.cell(row=epoch_now+1, column=1, value=epoch_now)  # 编号列
    ws.cell(row=epoch_now+1, column=2, value=str(fpr))  # 浮点数列，转换为字符串

    # 保存最佳模型
    if best_rf is None or best_fpr_lbf < best_fpr:
        best_rf = copy.deepcopy(rf)
        best_threshold = best_thresh
        best_fpr = best_fpr_lbf

    rf.n_estimators += 1  # 每一轮增加1

end_time = time.perf_counter_ns()
print(f'use {(end_time - start_time) / 1000000}ms')
wb.save('exp3_url_rf_lbf')
