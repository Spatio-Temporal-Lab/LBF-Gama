import copy
import time

import numpy as np
import pandas as pd
from openpyxl import Workbook
import torch
import torch.nn as nn
import torch.optim as optim

import lib.bf_util
import lib.lgb_url

# Load data
df_train = pd.read_csv('../dataset/url_train.csv')
df_test = pd.read_csv('../dataset/url_test.csv')
df_query = pd.read_csv('../dataset/url_query.csv')

train_urls = df_train['url']
test_urls = df_test['url']
query_urls = df_query['url']

X_train = df_train.drop(columns=['url', 'url_type']).values.astype(np.float32)
y_train = df_train['url_type'].values.astype(np.float32)
X_test = df_test.drop(columns=['url', 'url_type']).values.astype(np.float32)
y_test = df_test['url_type'].values.astype(np.float32)
X_query = df_query.drop(columns=['url', 'url_type']).values.astype(np.float32)
y_query = df_query['url_type'].values.astype(np.float32)

# Convert data to PyTorch tensors
X_train_tensor = torch.tensor(X_train)
y_train_tensor = torch.tensor(y_train).unsqueeze(1)  # Add an extra dimension for the output
X_test_tensor = torch.tensor(X_test)
y_test_tensor = torch.tensor(y_test).unsqueeze(1)  # Add an extra dimension for the output
X_query_tensor = torch.tensor(X_query)


# Define the neural network model
class ANNModel(nn.Module):
    def __init__(self, input_dim, hidden_dim):
        super(ANNModel, self).__init__()
        self.layer1 = nn.Linear(input_dim, hidden_dim)
        self.layer2 = nn.Linear(hidden_dim, 1)
        self.sigmoid = nn.Sigmoid()

    def forward(self, x):
        # x = torch.relu(self.layer1(x))
        x = self.sigmoid(self.layer1(x))
        x = self.sigmoid(self.layer2(x))
        return x


# Helper functions
def evaluate_thresholds(prediction_results, y_true, bf_bytes):
    sorted_indices = np.argsort(prediction_results)
    sorted_predictions = prediction_results[sorted_indices]
    sorted_true = y_true[sorted_indices]

    fp = n_false
    tp = 0
    best_thresh = 0
    best_fpr_lbf = 1.0

    unique_sorted_predictions, idx = np.unique(sorted_predictions, return_index=True)

    n = len(unique_sorted_predictions)
    for i in range(n):
        thresh = unique_sorted_predictions[i]

        if i < n - 1:
            count_1 = np.sum(sorted_true[idx[i]:idx[i + 1]])
            tp += count_1
            fp -= idx[i + 1] - idx[i] - count_1
        else:
            count_1 = np.sum(sorted_true[idx[i]:n])
            tp += count_1
            fp -= n - idx[i] - count_1

        bf_count = tp
        fpr_bf = lib.bf_util.get_fpr(bf_count, bf_bytes)
        fpr_lgb = fp / n_false
        fpr_lbf = fpr_lgb + (1 - fpr_lgb) * fpr_bf

        if fpr_lbf < best_fpr_lbf:
            best_thresh = thresh
            best_fpr_lbf = fpr_lbf

    print(f'best thresh = {best_thresh} and best fpr = {best_fpr_lbf}')
    return best_thresh, best_fpr_lbf


def ann_validate_url(model, X_train, y_train, train_urls, X_test, y_test, test_urls, threshold):
    model.eval()
    with torch.no_grad():
        prediction_results = model(X_train).squeeze().numpy()
    binary_predictions = (prediction_results > threshold).astype(int)
    indices_train = [i for i in range(len(X_train)) if binary_predictions[i] == 0 and y_train[i] == 1]

    with torch.no_grad():
        prediction_results = model(X_test).squeeze().numpy()
    binary_predictions = (prediction_results > threshold).astype(int)
    indices_test = [i for i in range(len(X_test)) if binary_predictions[i] == 0 and y_test[i] == 1]

    urls_train = train_urls.iloc[indices_train].values
    urls_test = test_urls.iloc[indices_test].values

    return np.concatenate((urls_train, urls_test), axis=0)


def ann_query_url(model, bloom_filter, X_query, y_query, query_urls, threshold, draw):
    fn = 0
    fp = 0
    cnt_ml = 0
    cnt_bf = 0
    total = len(X_query)
    print(f"query count = {total}")

    model.eval()
    with torch.no_grad():
        prediction_results = model(X_query).squeeze().numpy()

    if draw:
        import matplotlib.pyplot as plt
        bins = np.arange(0, 1.1, 0.1)
        plt.hist(prediction_results, bins=bins, edgecolor='black', alpha=0.7)
        plt.xlabel('Prediction value')
        plt.ylabel('Frequency')
        plt.title('Histogram of All Predictions')
        plt.xticks(bins)
        plt.grid(axis='y', alpha=0.75)
        plt.show()

    binary_predictions = (prediction_results > threshold).astype(int)

    print(prediction_results)
    print(binary_predictions)

    for i in range(total):
        true_label = y_query[i]
        url = query_urls[i]
        prediction = binary_predictions[i]
        if prediction == 1:
            if true_label == 0:
                fp = fp + 1
                cnt_ml = cnt_ml + 1
        else:
            if url in bloom_filter:
                if true_label == 0:
                    fp = fp + 1
                    cnt_bf = cnt_bf + 1
            else:
                if true_label == 1:
                    fn = fn + 1

    print(f"fp: {fp}")
    print(f"total: {total}")
    print(f"fpr: {float(fp) / total}")
    print(f"fnr: {float(fn) / total}")
    print(f"cnt_ml: {cnt_ml}")
    print(f"cnt_bf: {cnt_bf}")
    return float(fp) / total


# Training and evaluation
start_time = time.perf_counter_ns()

best_ann = None
size = 200 * 1024
best_fpr = 1.0
best_threshold = 0.5
epoch_max = 50  # Training each ANN for 50 epochs
wb = Workbook()
ws = wb.active

n_false = df_train[df_train['url_type'] == 0].shape[0] + df_test[df_test['url_type'] == 0].shape[0]
n_true = df_train[df_train['url_type'] == 1].shape[0] + df_test[df_test['url_type'] == 1].shape[0]
n_test = len(df_test)
input_dim = X_train.shape[1]

for hidden_dim in range(10, 201, 10):  # Outer loop for number of neurons in hidden layer
    model = ANNModel(input_dim, hidden_dim)
    criterion = nn.BCELoss()
    optimizer = optim.Adam(model.parameters(), lr=0.001)

    for epoch_now in range(epoch_max):
        model.train()
        optimizer.zero_grad()
        outputs = model(X_train_tensor)
        loss = criterion(outputs, y_train_tensor)
        loss.backward()
        optimizer.step()

    model.eval()
    with torch.no_grad():
        train_pred = model(X_train_tensor).squeeze().numpy()
        test_pred = model(X_test_tensor).squeeze().numpy()

    all_predictions = np.concatenate([train_pred, test_pred])
    all_true_labels = np.concatenate([y_train, y_test])

    model_size = lib.lgb_url.lgb_get_model_size(model)
    print(f'Hidden neurons = {hidden_dim}')
    print(f'Model size = {model_size}')
    bf_bytes = size - model_size
    if bf_bytes <= 0:
        break

    print(f'all_predictions: {all_predictions}')
    best_thresh, best_fpr_lbf = evaluate_thresholds(all_predictions, all_true_labels, bf_bytes)

    print("模型在内存中所占用的大小（字节）:", model_size)

    data_negative = ann_validate_url(model, X_train_tensor, y_train, train_urls, X_test_tensor, y_test, test_urls,
                                     best_thresh)
    print(f"{len(data_negative)} insert into bloom filter")
    bloom_filter = lib.lgb_url.create_bloom_filter(dataset=data_negative, bf_size=bf_bytes)
    num_bits = bloom_filter.num_bits
    memory_in_bytes = num_bits / 8
    print("memory of bloom filter: ", memory_in_bytes)
    print("memory of learned model: ", model_size)
    fpr = ann_query_url(model, bloom_filter, X_query_tensor, y_query, query_urls, best_thresh, False)

    row = hidden_dim // 10
    ws.cell(row=row, column=1, value=hidden_dim)
    ws.cell(row=row, column=2, value=str(fpr))

    if best_ann is None or best_fpr_lbf < best_fpr:
        best_ann = copy.deepcopy(model)
        best_threshold = best_thresh
        best_fpr = best_fpr_lbf

end_time = time.perf_counter_ns()
print(f'use {(end_time - start_time) / 1000000}ms')
wb.save('results/exp3_url_ann_lbf.xlsx')