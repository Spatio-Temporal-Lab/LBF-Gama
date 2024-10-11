import copy
import time

import numpy as np
import pandas as pd
import torch
import torch.nn as nn
import torch.optim as optim
from openpyxl import Workbook

import lib.lgb_url
from plbf.FastPLBF_M import FastPLBF_M

# Load data
df_train = pd.read_csv('../dataset/url_train.csv')
df_test = pd.read_csv('../dataset/url_test.csv')
df_query = pd.read_csv('../dataset/url_query.csv')

train_urls = df_train['url']
test_urls = df_test['url']
query_urls = df_query['url']
positive_train_urls = df_train[df_train['url_type'] == 1]['url']
positive_test_urls = df_test[df_test['url_type'] == 1]['url']
positive_urls = pd.concat([positive_train_urls, positive_test_urls])
positive_urls_list = positive_urls.tolist()

X_train = df_train.drop(columns=['url', 'url_type']).values.astype(np.float32)
y_train = df_train['url_type'].values.astype(np.float32)
X_test = df_test.drop(columns=['url', 'url_type']).values.astype(np.float32)
y_test = df_test['url_type'].values.astype(np.float32)
X_query = df_query.drop(columns=['url', 'url_type']).values.astype(np.float32)
y_query = df_query['url_type'].values.astype(np.float32)
positive_samples = np.concatenate((X_train[y_train == 1], X_test[y_test == 1]), axis=0)
negative_samples = np.concatenate((X_train[y_train == 0], X_test[y_test == 0]), axis=0)

# Convert data to PyTorch tensors
X_train_tensor = torch.tensor(X_train)
y_train_tensor = torch.tensor(y_train).unsqueeze(1)  # Add an extra dimension for the output
X_test_tensor = torch.tensor(X_test)
y_test_tensor = torch.tensor(y_test).unsqueeze(1)  # Add an extra dimension for the output
X_query_tensor = torch.tensor(X_query)
positive_tensor = torch.tensor(positive_samples)
negative_tensor = torch.tensor(negative_samples)


# Define the neural network model
class ANNModel(nn.Module):
    def __init__(self, input_dim, hidden_dim):
        super(ANNModel, self).__init__()
        self.layer1 = nn.Linear(input_dim, hidden_dim)
        self.layer2 = nn.Linear(hidden_dim, 1)
        self.sigmoid = nn.Sigmoid()

    def forward(self, x):
        # x = torch.relu(self.layer1(x))
        x = self.sigmoid(self.layer1(x))
        x = self.sigmoid(self.layer2(x))
        return x


# Training and evaluation
start_time = time.perf_counter_ns()

best_ann = None
size = 200 * 1024
best_fpr = 1.0
best_threshold = 0.5
epoch_max = 50  # Training each ANN for 50 epochs
wb = Workbook()
ws = wb.active

n_false = df_train[df_train['url_type'] == 0].shape[0] + df_test[df_test['url_type'] == 0].shape[0]
n_true = df_train[df_train['url_type'] == 1].shape[0] + df_test[df_test['url_type'] == 1].shape[0]
n_test = len(df_test)
input_dim = X_train.shape[1]

for hidden_dim in range(20, 201, 10):  # Outer loop for number of neurons in hidden layer
    model = ANNModel(input_dim, hidden_dim)
    criterion = nn.BCELoss()
    optimizer = optim.Adam(model.parameters(), lr=0.001)

    for epoch_now in range(epoch_max):
        model.train()
        optimizer.zero_grad()
        outputs = model(X_train_tensor)
        loss = criterion(outputs, y_train_tensor)
        loss.backward()
        optimizer.step()

    model_size = lib.lgb_url.lgb_get_model_size(model)
    print(f'Hidden neurons = {hidden_dim}')
    print(f'Model size = {model_size}')
    bf_bytes = size - model_size
    if bf_bytes <= 0:
        break

    model.eval()
    with torch.no_grad():
        pos_scores = model(positive_tensor).squeeze().numpy().tolist()
        neg_scores = model(negative_tensor).squeeze().numpy().tolist()

    plbf = FastPLBF_M(positive_urls_list, pos_scores, neg_scores, bf_bytes * 8.0, 50, 5)
    plbf.insert_keys(positive_urls_list, pos_scores)

    fp_cnt = 0
    query_negative = X_query
    query_neg_keys = query_urls
    model.eval()
    with torch.no_grad():
        query_neg_scores = model(X_query_tensor).squeeze().numpy()
    total = len(query_negative)
    for key, score in zip(query_neg_keys, query_neg_scores):
        if plbf.contains(key, score):
            fp_cnt += 1
    fpr_test = float(fp_cnt) / total
    print(f"test fpr: {fpr_test}")

    row = hidden_dim // 10
    ws.cell(row=row, column=1, value=hidden_dim)
    ws.cell(row=row, column=2, value=str(fpr_test))

end_time = time.perf_counter_ns()
print(f'use {(end_time - start_time) / 1000000}ms')
wb.save('results/exp3_url_ann_plbf.xlsx')
