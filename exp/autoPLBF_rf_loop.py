import copy
import time

import numpy as np
import pandas as pd
from openpyxl import Workbook
from sklearn.ensemble import RandomForestClassifier

import lib.lgb_url
from plbf.FastPLBF_M import FastPLBF_M

df_train = pd.read_csv('../dataset/url_train.csv')
df_test = pd.read_csv('../dataset/url_test.csv')
df_query = pd.read_csv('../dataset/url_query.csv')

train_urls = df_train['url']
test_urls = df_test['url']
query_urls = df_query['url']

positive_train_urls = df_train[df_train['url_type'] == 1]['url']
positive_test_urls = df_test[df_test['url_type'] == 1]['url']
positive_urls = pd.concat([positive_train_urls, positive_test_urls])
positive_urls_list = positive_urls.tolist()

X_train = df_train.drop(columns=['url', 'url_type']).values.astype(np.float32)
y_train = df_train['url_type'].values.astype(np.float32)
X_test = df_test.drop(columns=['url', 'url_type']).values.astype(np.float32)
y_test = df_test['url_type'].values.astype(np.float32)
X_query = df_query.drop(columns=['url', 'url_type']).values.astype(np.float32)
y_query = df_query['url_type'].values.astype(np.float32)
positive_samples = np.concatenate((X_train[y_train == 1], X_test[y_test == 1]), axis=0)
negative_samples = np.concatenate((X_train[y_train == 0], X_test[y_test == 0]), axis=0)

# 设置参数
params = {
    'n_estimators': 1,  # 每一轮训练的树的数量
    'random_state': 42,
    'max_leaf_nodes': 20,
    'warm_start': True,  # 启用 warm_start 以便在下一轮中继续训练
}

n_true = df_train[df_train['url_type'] == 1].shape[0] + df_test[df_test['url_type'] == 1].shape[0]
n_false = df_train[df_train['url_type'] == 0].shape[0] + df_test[df_test['url_type'] == 0].shape[0]
n_test = len(df_test)


start_time = time.perf_counter_ns()

best_rf = None
size = 200 * 1024
best_fpr = 1.0
best_threshold = 0.5
epoch_max = 20
wb = Workbook()
ws = wb.active
rf = RandomForestClassifier(**params)

for epoch_now in range(epoch_max):
    rf.fit(X_train, y_train)
    model_size = lib.lgb_url.lgb_get_model_size(rf)
    bf_size = size - model_size
    print(f'size: {model_size}, {bf_size}')
    if bf_size <= 0:
        break
    # if epoch_now < 1:
    #     continue

    pos_scores = rf.predict_proba(positive_samples)[:, 1].tolist()
    neg_scores = rf.predict_proba(negative_samples)[:, 1].tolist()

    plbf = FastPLBF_M(positive_urls_list, pos_scores, neg_scores, bf_size * 8.0, 50, 5)
    plbf.insert_keys(positive_urls_list, pos_scores)
    fpr = plbf.get_fpr()

    fp_cnt = 0
    query_negative = X_query
    query_neg_keys = query_urls
    query_neg_scores = rf.predict_proba(X_query)[:, 1]
    total = len(query_negative)

    for key, score in zip(query_neg_keys, query_neg_scores):
        if plbf.contains(key, score):
            fp_cnt += 1
    fpr_test = float(fp_cnt) / total
    print(f"test fpr: {fpr_test}")
    ws.cell(row=epoch_now + 1, column=1, value=epoch_now + 1)  # 编号列
    ws.cell(row=epoch_now + 1, column=2, value=str(fpr_test))  # 浮点数列，转换为字符串

    if best_rf is None or fpr < best_fpr:
        best_bst = copy.deepcopy(rf)
        best_fpr = fpr
        best_epoch = epoch_now
        best_plbf = plbf

    rf.n_estimators += 1  # 每一轮增加1

end_time = time.perf_counter_ns()
print(f'use {(end_time - start_time) / 1000000}ms')
wb.save('results/exp3_url_rf_plbf.xlsx')
